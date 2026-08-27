"""Excel extraction, document-style. `M1-EXTRACT-ING-027`.

**A document-style read, not the queryable table `M1-DATA-*` in M4 builds.**
Every non-empty row of every sheet becomes one anchor, text joined cell by
cell — enough to cite "which row said this" without pretending it is a typed,
queryable table. That arrives in M4, against `docs/data-sources.md` §2.

**`openpyxl`, MIT-licensed, `read_only=True`.** A workbook here can be
whatever size somebody's spreadsheet habit produced; read-only mode streams
rows rather than materialising the whole sheet.

**Multi-sheet and merged cells: handled, crudely, and said so out loud.**
Every sheet is read — nothing here special-cases which one is "the" data —
and a merged range's value lives only in its top-left cell in the underlying
XML, which `openpyxl` already surfaces as `None` for the rest of the range
without this module doing anything about it. `docs/data-sources.md` §8 records
the real fix (ask, don't guess) as out of this ticket's scope; the merged
count is logged so the gap is visible rather than silent.
"""

import asyncio
from typing import TYPE_CHECKING, cast

import openpyxl
from openpyxl.workbook.workbook import Workbook

from askwell.extract_common import Anchor, write_anchors
from askwell.logging import get_logger

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

    from askwell.ingest import Report, Work

log = get_logger(__name__)

ANCHOR_KIND = "sheet_row"


def _row_text(values: tuple[object, ...]) -> str:
    cells = [str(value).strip() for value in values if value is not None and str(value).strip()]
    return " | ".join(cells)


def _load(path: str) -> Workbook:  # type: ignore[no-any-unimported]
    return cast(  # type: ignore[no-any-unimported]
        Workbook, openpyxl.load_workbook(path, data_only=True, read_only=True)
    )


def _read(path: str) -> tuple[list[tuple[str, str]], int, int]:
    """The blocking half: open the workbook and walk every sheet.

    Run through `asyncio.to_thread` as one call rather than per row — unlike
    `extract_pdf`'s per-page loop, `openpyxl`'s read-only iterator is itself
    the streaming boundary, and yielding control to the event loop mid-sheet
    would mean holding the file handle open across awaits for no benefit.
    """
    workbook = _load(path)
    try:
        rows: list[tuple[str, str]] = []
        merged_total = 0
        for name in workbook.sheetnames:
            worksheet = workbook[name]
            # `read_only` worksheets carry no merge information at all —
            # `openpyxl` only tracks it on the fully-loaded kind. The count is
            # informational (logged, not acted on), so a sheet read-only just
            # contributes nothing to it rather than forcing a second, slower
            # non-read-only pass purely to count merges.
            merges = getattr(worksheet, "merged_cells", None)
            if merges is not None:
                merged_total += len(merges.ranges)
            for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
                text = _row_text(values)
                if text:
                    rows.append((f"{name}, row {row_number}", text))
        return rows, len(workbook.sheetnames), merged_total
    finally:
        workbook.close()


async def run(work: "Work", report: "Report", factory: "async_sessionmaker[AsyncSession]") -> None:
    rows, sheet_count, merged_total = await asyncio.to_thread(_read, work.path)

    anchors: list[Anchor] = []
    for index, (label, text) in enumerate(rows, start=1):
        anchors.append(Anchor(page_number=index, label=label, text=text, has_text=True))
        await report(index, len(rows))

    await write_anchors(factory, work, anchors, ANCHOR_KIND)

    log.info(
        "extract_xlsx_completed",
        document_id=str(work.document_id),
        filename=work.filename,
        sheets=sheet_count,
        rows=len(rows),
        merged_ranges=merged_total,
    )
